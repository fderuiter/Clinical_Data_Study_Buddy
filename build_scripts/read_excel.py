import openpyxl
import sys

def read_excel(filepath):
    try:
        workbook = openpyxl.load_workbook(filepath)
        for sheet_name in workbook.sheetnames:
            print(f"--- Sheet: {sheet_name} ---")
            worksheet = workbook[sheet_name]
            for row in worksheet.iter_rows():
                print([cell.value for cell in row])
    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        sys.exit(1)

if __name__ == "__main__":
    if len(sys.argv) > 1:
        read_excel(sys.argv[1])
    else:
        print("Usage: python read_excel.py <path_to_excel_file>")
