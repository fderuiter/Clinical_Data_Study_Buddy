import os
import pytest
from unittest.mock import MagicMock, patch
import pandas as pd
import openpyxl
from pathlib import Path
import argparse

from scripts.generate_spec_template import generate_spec_template
from scripts.generate_dataset_from_spec import generate_dataset_from_spec
from scripts.validate_dataset import validate_dataset


@pytest.fixture
def mock_cdisc_library_client():
    with patch(
        "scripts.generate_spec_template.get_mdr_sdtmig_version_datasets_dataset.sync"
    ) as mock_sync:
        mock_response = MagicMock()
        mock_response.to_dict.return_value = {
            "datasetVariables": [
                {
                    "name": "STUDYID",
                    "label": "Study Identifier",
                    "simpleDatatype": "Char",
                    "role": "Identifier",
                    "core": "Req",
                    "_links": {"codelist": [{"href": "/a/b/c"}]},
                    "description": "Unique identifier for a study.",
                }
            ]
        }
        mock_sync.return_value = mock_response
        yield mock_sync


def test_generate_spec_template(mock_cdisc_library_client, tmp_path):
    output_dir = tmp_path
    with patch("scripts.generate_spec_template.get_client"):
        generate_spec_template("sdtmig", "3-3", ["DM"], str(output_dir))
        output_file = output_dir / "sdtmig_3-3_spec.xlsx"
        assert output_file.exists()

        workbook = openpyxl.load_workbook(output_file)
        assert "DM" in workbook.sheetnames
        worksheet = workbook["DM"]
        header = [cell.value for cell in worksheet[1]]
        assert header == [
            "Variable Name",
            "Variable Label",
            "Data Type",
            "Role",
            "Core",
            "Codelist",
            "Description",
        ]
        first_row = [cell.value for cell in worksheet[2]]
        assert first_row[0] == "STUDYID"


@pytest.fixture
def mock_dataset_generator_client():
    with patch(
        "scripts.generate_dataset_from_spec.CDISCDataSetGeneratorClient"
    ) as mock_client:
        instance = mock_client.return_value
        instance.generate_dataset.return_value = {
            "download_url": "http://example.com/data.csv",
            "filename": "sdtm_dm_test.csv",
        }
        instance.download_file.return_value = None
        yield instance


def test_generate_dataset_from_spec(mock_dataset_generator_client, tmp_path):
    # Create a dummy spec file
    spec_file = tmp_path / "sdtmig_3-3_spec.xlsx"
    workbook = openpyxl.Workbook()
    # Remove the default sheet
    workbook.remove(workbook.active)
    workbook.create_sheet("DM")
    workbook.save(spec_file)

    generate_dataset_from_spec(str(spec_file), str(tmp_path))

    mock_dataset_generator_client.generate_dataset.assert_called_with(
        dataset_type="SDTM",
        domain="DM",
        num_subjects=50,
        therapeutic_area="Oncology",
        format="csv",
    )
    mock_dataset_generator_client.download_file.assert_called_once()


@pytest.fixture
def sample_spec_and_data(tmp_path):
    # Create a spec file
    spec_file = tmp_path / "sdtmig_3-3_spec.xlsx"
    spec_df = pd.DataFrame(
        {
            "Variable Name": ["STUDYID", "USUBJID", "AGE"],
            "Data Type": ["Char", "Char", "Num"],
        }
    )
    # Use a writer to save to a specific sheet
    with pd.ExcelWriter(spec_file, engine='openpyxl') as writer:
        spec_df.to_excel(writer, sheet_name="DM", index=False)


    # Create a valid data file
    valid_data_file = tmp_path / "sdtm_dm_valid.csv"
    valid_df = pd.DataFrame(
        {"STUDYID": ["101"], "USUBJID": ["101-001"], "AGE": [34]}
    )
    valid_df.to_csv(valid_data_file, index=False)

    # Create an invalid data file
    invalid_data_file = tmp_path / "sdtm_dm_invalid.csv"
    invalid_df = pd.DataFrame(
        {"STUDYID": ["101"], "AGE": ["thirty-four"], "EXTRA_COL": [1]}
    )
    invalid_df.to_csv(invalid_data_file, index=False)

    return spec_file, valid_data_file, invalid_data_file


def test_validate_dataset_success(sample_spec_and_data, capsys):
    spec_file, valid_data_file, _ = sample_spec_and_data
    validate_dataset(str(spec_file), str(valid_data_file))
    captured = capsys.readouterr()
    assert "Validation Successful" in captured.out


def test_validate_dataset_failure(sample_spec_and_data, capsys):
    spec_file, _, invalid_data_file = sample_spec_and_data
    validate_dataset(str(spec_file), str(invalid_data_file))
    captured = capsys.readouterr()
    assert "Validation Failed" in captured.out
    assert "Missing columns" in captured.out
    assert "Extra columns" in captured.out
    assert "Data type error" in captured.out
