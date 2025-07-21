import os
import subprocess
from pathlib import Path

from crfgen.schema import load_forms
from crfgen.exporter import EXPORTERS
import crfgen.exporter.xlsx  # register xlsx exporter


def test_render_xlsx(tmp_path: Path):
    forms = load_forms("tests/.data/sample_crf.json")
    exporter = EXPORTERS["xlsx"]
    exporter(forms, tmp_path)
    xlsx_file = tmp_path / "forms.xlsx"
    assert xlsx_file.exists()
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_file)
    assert wb.sheetnames
    sheet = wb[wb.sheetnames[0]]
    header = [cell.value for cell in next(sheet.iter_rows(values_only=False))]
    assert header[:4] == ["OID", "Prompt", "Datatype", "Codelist"]


def test_build_script(tmp_path: Path):
    cmd = [
        "python",
        "scripts/build.py",
        "--source",
        "tests/.data/sample_crf.json",
        "--outdir",
        str(tmp_path),
        "--formats",
        "xlsx",
    ]
    env = os.environ.copy()
    env["PYTHONPATH"] = "src"
    subprocess.check_call(cmd, env=env)
    assert (Path(tmp_path) / "forms.xlsx").exists()
