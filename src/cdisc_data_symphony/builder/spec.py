"""
This module provides functionalities for working with CDISC dataset specifications.

It includes functions to:
- Generate an Excel-based specification template from the CDISC Library.
- Generate a synthetic dataset from a specification template.
- Validate a dataset against its corresponding specification template.
"""
import os
from pathlib import Path
import json
import openpyxl
import pandas as pd
from cdisc_library_client.client import AuthenticatedClient
from cdisc_library_client.api.sdtm_implementation_guide_sdtmig import (
    get_mdr_sdtmig_version_datasets_dataset,
)
from cdisc_library_client.api.analysis_data_model_and_implementation_guide_a_da_m_and_a_da_mig import (
    get_mdr_adam_product_datastructures_structure,
)
from dotenv import load_dotenv
from cdisc_data_symphony.builder.data_generator import DataGenerator
from cdisc_data_symphony.builder.crfgen.schema import Form, FieldDef


def get_client():
    """
    Creates and returns an authenticated client for the CDISC Library API.

    This function loads environment variables, retrieves the API key, and
    initializes an AuthenticatedClient.

    Returns:
        AuthenticatedClient: An authenticated client for the CDISC Library API.

    Raises:
        ValueError: If the CDISC Library API key is not found in the environment variables.
    """
    load_dotenv()
    api_key = os.environ.get("CDISC_PRIMARY_KEY")
    if not api_key:
        raise ValueError(
            "CDISC Library API key not found. Please set the CDISC_PRIMARY_KEY environment variable."
        )
    headers = {"api-key": api_key}
    return AuthenticatedClient(
        base_url="https://library.cdisc.org/api", headers=headers, token="dummy"
    )

def generate_template(
    product: str, version: str, domains: list[str], output_dir: str
):
    """
    Generates an Excel-based specification template for CDISC datasets.

    This function fetches metadata from the CDISC Library for the specified product,
    version, and domains, and then creates an Excel spreadsheet with the specification.

    Args:
        product (str): The CDISC product (e.g., "sdtmig", "adamig").
        version (str): The version of the product (e.g., "3-3").
        domains (list[str]): A list of domains to include in the specification.
        output_dir (str): The directory where the generated Excel file will be saved.
    """
    client = get_client()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    print(f"Generating spec for {product} v{version} for domains: {', '.join(domains)}")

    for domain in domains:
        print(f"Fetching data for domain: {domain}")
        worksheet = workbook.create_sheet(title=domain)
        header = [
            "Variable Name",
            "Variable Label",
            "Data Type",
            "Role",
            "Core",
            "Codelist",
            "Description",
        ]
        worksheet.append(header)
        try:
            variables = []
            if product.lower() == "sdtmig":
                response = get_mdr_sdtmig_version_datasets_dataset.sync(
                    client=client, version=version, dataset=domain
                )
                if response and hasattr(response, "to_dict"):
                    dataset = response.to_dict()
                    variables = dataset.get("datasetVariables", [])
            elif product.lower() == "adamig":
                adam_product_id = f"adamig-{version}"
                response = get_mdr_adam_product_datastructures_structure.sync(
                    client=client, product=adam_product_id, structure=domain
                )
                if response and hasattr(response, "to_dict"):
                    dataset = response.to_dict()
                    varsets = dataset.get("analysisVariableSets", [])
                    for varset in varsets:
                        variables.extend(varset.get("analysisVariables", []))

            if not variables:
                print(f"No variables found for domain {domain}")
                worksheet.append([f"No variables found for domain {domain}"])
                continue

            for variable in variables:
                codelist_info = ""
                if variable.get("_links", {}).get("codelist"):
                    codelists = variable["_links"]["codelist"]
                    codelist_info = ", ".join([c.get("href", "") for c in codelists])
                row = [
                    variable.get("name"),
                    variable.get("label"),
                    variable.get("simpleDatatype"),
                    variable.get("role"),
                    variable.get("core"),
                    codelist_info,
                    variable.get("description"),
                ]
                worksheet.append(row)

        except Exception as e:
            print(f"Could not fetch data for domain {domain}: {e}")
            worksheet.append([f"Error fetching data for {domain}"])

    output_path = Path(output_dir) / f"{product}_{version}_spec.xlsx"
    workbook.save(output_path)
    print(f"Specification template generated at: {output_path}")


def generate_dataset(spec_path: str, output_dir: str):
    """
    Generates a synthetic dataset from a specification template.

    This function reads an Excel-based specification template, and for each domain
    (sheet) in the template, it generates a synthetic CSV dataset.

    Args:
        spec_path (str): The path to the Excel specification template.
        output_dir (str): The directory where the generated CSV files will be saved.
    """
    path = Path(spec_path)
    workbook = openpyxl.load_workbook(path)
    domains = workbook.sheetnames

    for domain in domains:
        if domain.lower() == "metadata":
            continue
        print(f"Generating dataset for domain {domain}...")

        spec_df = pd.read_excel(path, sheet_name=domain)
        fields = []
        for _, row in spec_df.iterrows():
            fields.append(
                FieldDef(
                    oid=row["Variable Name"],
                    prompt=row["Variable Label"],
                    datatype=row["Data Type"],
                    cdash_var=row["Variable Name"],
                )
            )

        form_data = Form(title=domain, domain=domain, fields=fields)
        generator = DataGenerator(form_data)
        dataset = generator.generate(num_subjects=50)

        output_path = Path(output_dir) / f"{domain}.csv"
        df = pd.DataFrame(dataset)
        df.to_csv(output_path, index=False)
        print(f"Dataset for domain {domain} generated successfully at {output_path}")

def validate(spec_path: str, dataset_path: str):
    """
    Validates a dataset against a specification template.

    This function checks a given dataset (in CSV format) against an Excel-based
    specification template to ensure that the dataset conforms to the spec.

    Args:
        spec_path (str): The path to the Excel specification template.
        dataset_path (str): The path to the CSV dataset to be validated.
    """
    spec_path = Path(spec_path)
    dataset_path = Path(dataset_path)

    dataset_stem = dataset_path.stem
    parts = dataset_stem.split("_")
    if len(parts) == 3:  # product_domain_ts
        domain = parts[1].upper()
    elif len(parts) == 1:  # domain
        domain = parts[0].upper()
    else:
        print(
            f"Invalid dataset filename format: {dataset_path.name}. Expected '<product>_<domain>_<timestamp>.csv' or '<domain>.csv'."
        )
        return

    print(f"Validating dataset {dataset_path.name} against spec {spec_path.name}")

    try:
        spec_df = pd.read_excel(spec_path, sheet_name=domain)
    except ValueError:
        print(f"Sheet '{domain}' not found in the specification file.")
        return

    dataset_df = pd.read_csv(dataset_path, dtype=str)

    validation_warnings = []
    validation_errors = []

    spec_columns = spec_df["Variable Name"].tolist()
    dataset_columns = dataset_df.columns.tolist()
    missing_columns = set(spec_columns) - set(dataset_columns)
    if missing_columns:
        validation_warnings.append(f"Missing columns in dataset that are in the spec: {', '.join(missing_columns)}")

    extra_columns = set(dataset_columns) - set(spec_columns)
    if extra_columns:
        validation_errors.append(f"Extra columns in dataset that are not in the spec: {', '.join(extra_columns)}")

    for _, row in spec_df.iterrows():
        col_name = row["Variable Name"]
        spec_type = row["Data Type"]
        if col_name in dataset_df.columns:
            if spec_type == "Num":
                try:
                    pd.to_numeric(dataset_df[col_name].dropna())
                except (ValueError, TypeError):
                    validation_errors.append(f"Data type error in column '{col_name}': Expected a numeric type.")

    if validation_errors:
        print("\nValidation Failed:")
        for error in validation_errors:
            print(f"- {error}")
    else:
        print("\nValidation Successful: Dataset conforms to the specification.")

    if validation_warnings:
        print("\nValidation Warnings:")
        for warning in validation_warnings:
            print(f"- {warning}")
